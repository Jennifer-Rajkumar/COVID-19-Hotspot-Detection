# -*- coding: utf-8 -*-
"""
Created on Wed Mar 17 00:08:41 2021

@author: Jennifer
"""

import openpyxl 
import pymongo
  

wrkbk = openpyxl.load_workbook("C:/Users/Jennifer/.spyder-py3/Hotspot1/Hotspot/vaccine.xlsx") 
  
sh = wrkbk.active 

for row in sh.iter_rows(min_row=1, min_col=1, max_row=7936, max_col=14): 
    for cell in row: 
        print(cell.value, end=" ") 
    print()

myclient = pymongo.MongoClient("mongodb://localhost:27017/")
mydb = myclient["vaccine"]
mycol = mydb["hospital"]
dic={}

for row in sh.iter_rows(): 
    l=[]
    for cell in row:
        l.append(cell.value)
    link='https://www.google.com/maps/place/'+str(l[4])+','+str(l[5])
    if l[6].lower() not in dic:
        dic[l[6].lower()]={l[7].lower() : [[str(l[0]),str(l[1]),str(l[2]),str(l[3]),link,str(l[6]),str(l[7]),str(l[9]),str(l[10]),str(l[11]),str(l[12])]]}
    else:
        if l[7].lower() not in dic[l[6].lower()]:
            dic[l[6].lower()][l[7].lower()]=[[str(l[0]),str(l[1]),str(l[2]),str(l[3]),link,str(l[6]),str(l[7]),str(l[9]),str(l[10]),str(l[11]),str(l[12])]]
        else:
            dic[l[6].lower()][l[7].lower()].append([str(l[0]),str(l[1]),str(l[2]),str(l[3]),link,str(l[6]),str(l[7]),str(l[9]),str(l[10]),str(l[11]),str(l[12])])

for k,v in dic.items():
    for x,y in v.items():
        mycol.insert_one({"state":k,"district":x, "details":y})
                          